from operator import index
from openpyxl import Workbook
import datetime

planilha = Workbook()
ws = planilha.active
ws['A1'] = 'ID'
ws['B1'] = 'Nome'
ws['C1'] = 'Hora'
a=planilha['Sheet']
index = 2
hora = datetime.datetime.now()
id = [1, 2, 3, 89]
nome = ['Kelven', 'Lorena', 'Myara', 'teste']

for id, nome in zip(id, nome):
    a.cell(column=1, row=index, value=id)
    #nome = input('Digite o nome')
    a.cell(column=2, row=index, value=nome)
    a.cell(column=3, row=index, value=hora)
    index +=1
planilha.save("kelventeste.xlsx")

print('salvo')