from time import sleep
import openpyxl
from selenium import webdriver
from operator import index
from openpyxl import Workbook
from selenium.webdriver.common.by import By

class web():
    def __init__(self) -> None:
        self.navegador = webdriver.Chrome()
        self.navegador.set_window_size(800,700)
        self.navegador.get('https://gatry.com/')
        print('voce esta no site: ', self.navegador.title) 
        self.lista_produto = []
        self.lista_preco = []

        self.pegando_nome_produto()
        self.pegando_preco_produto()
        self.criar_planilha()
        
        self.navegador.quit()

    def pegando_nome_produto(self):
        items = self.navegador.find_elements(By.XPATH, 
        '//div[@class="description"]//h3//a')
        for item in items:
            self.lista_produto.append(item.text)
            
    

    def pegando_preco_produto(self):
        items = self.navegador.find_elements(By.XPATH, 
        '//div[@class="description"]//p[@class="price"]')
        for item in items:
            self.lista_preco.append(item.text)

    def criar_planilha(self):
        index = 2
        planilha = openpyxl.Workbook()
        produtos = planilha['Sheet']
        produtos.title = 'Produtos'
        produtos['A1'] = 'PRODUTOS'
        produtos['B1'] = 'PRECOS'
        for nome, preco in zip(self.lista_produto, self.lista_preco):
            produtos.cell(column=1, row=index, value=nome)
            produtos.cell(column=2, row=index, value=preco)
            index+=1
        planilha.save('promocoes_gatry.xlsx')
        print(f'\u001b[32m{"Planilha criada com sucesso"}\u001b[0m')
        self.navegador.quit()


start = web()