from time import sleep
import openpyxl
from selenium import webdriver
from operator import index
from openpyxl import Workbook
from selenium.webdriver.common.by import By


class web():
    def __init__(self):
        
        self.abrir_navegador()
        #self.ver_preco()
        #self.criar_planilha()


    def abrir_navegador(self):
        self.navegador = webdriver.Chrome()
        self.navegador.set_window_size(800,700)
        self.navegador.get('https://telefonesimportados.netlify.app/')
        print('voce esta no site: ', self.navegador.title) 
        self.lista_produto = []
        self.lista_preco = []
        lista_nomes = self.navegador.find_element(by=By.XPATH, value='//div[@class="description"]//h3//a')
        self.lista_produto.append(lista_nomes.text)
        print(self.lista_produto)
        
    def ver_preco(self):
        item = 1
        
        for i in range (5):
            lista_nomes = self.navegador.find_element(by=By.XPATH, value='//div[@class="description"]//h3//a')
            self.lista_produto.append(lista_nomes.text)
            
            
            lista_preco = self.navegador.find_element(by=By.XPATH, value=f'/html/body/div[5]/div[2]/div[1]/div[{item}]/div/div[2]/ins')
            self.lista_preco.append(lista_preco.text)
            item +=1
        print(self.lista_preco)
        print(self.lista_produto)
    def criar_planilha(self):
        index = 2
        planilha = openpyxl.Workbook()
        produtos = planilha['Sheet']
        produtos.title = 'Produtos'
        produtos['A1'] = 'PRODUTOS'
        produtos['B1'] = 'PRECOS'
        for nome, preco in zip(self.lista_produto, self.lista_preco):
            produtos.cell(column=1, row=index, value=nome)
            produtos.cell(column=2, row=index, value=preco)
            index+=1
        planilha.save('promocoes_gatry.xlsx')
        print(f'\u001b[32m{"Planilha criada com sucesso"}\u001b[0m')
        self.navegador.quit()






start = web()

